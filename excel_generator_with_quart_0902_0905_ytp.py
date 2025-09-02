# 代码生成时间: 2025-09-02 09:05:42
import quart
from quart import request, jsonify
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook
from openpyxl.writer.excel import ExcelWriter
import os
from pathlib import Path

# 定义一个全局变量来保存生成的Excel文件的路径
BASE_DIR = Path().cwd() / 'generated_excels'

# 确保保存Excel文件的目录存在
BASE_DIR.mkdir(parents=True, exist_ok=True)

app = quart.Quart(__name__)

"""
该函数处理生成Excel表格的请求。
它接收一个包含表头的JSON数组，并生成一个Excel文件。
"""
@app.route('/generate-excel', methods=['POST'])
def generate_excel():
    try:
        # 获取请求中的JSON数据
        data = request.get_json()
        if not data or 'headers' not in data or 'rows' not in data:
            return jsonify({'error': 'Invalid data format'}), 400

        headers = data['headers']
        rows = data['rows']

        # 创建一个新的工作簿
        workbook = Workbook()
        sheet = workbook.active

        # 将表头写入Excel表格
        for col, header in enumerate(headers, start=1):
            sheet.cell(row=1, column=col).value = header

        # 将数据行写入Excel表格
        for row_index, row in enumerate(rows, start=2):
            for col_index, value in enumerate(row, start=1):
                sheet.cell(row=row_index, column=col_index).value = value

        # 获取文件名
        filename = f'generated_{quart.utils.get_debug_flag()}_{int(time.time())}.xlsx'
        file_path = BASE_DIR / filename

        # 保存Excel文件
        workbook.save(str(file_path))

        return jsonify({'message': f'Excel file generated successfully.',
                        'filename': filename}), 200
    except InvalidFileException as e:
        return jsonify({'error': str(e)}), 400
    except Exception as e:
        return jsonify({'error': 'An unexpected error occurred'}), 500

"""
启动Quart应用
"""
if __name__ == '__main__':
    app.run(debug=True)
